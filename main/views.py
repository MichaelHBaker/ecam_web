# views.py
from django.shortcuts import render, redirect
from django.db import models
from django.db.models import Prefetch, Q
from django.http import JsonResponse
from django.contrib.auth import authenticate, login
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.utils.timezone import get_fixed_timezone 
from django.views.decorators.http import require_http_methods
from django.urls import reverse

from zoneinfo import available_timezones

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.permissions import IsAuthenticated

from .models import (
    Project, Location, Measurement, MeasurementCategory,
    MeasurementType, MeasurementUnit, DataImport, DataSource,
    Dataset, DataSourceLocation, ProjectAccess, DataCopyGrant
)
from .serializers import (
    ProjectSerializer, LocationSerializer, MeasurementSerializer,
    MeasurementCategorySerializer, MeasurementTypeSerializer,
    MeasurementUnitSerializer, ModelFieldsSerializer,
    DataSourceSerializer, DataSourceLocationSerializer,
    DatasetSerializer, SourceColumnSerializer, DataImportSerializer,
    ImportBatchSerializer, ProjectAccessSerializer, DataCopyGrantSerializer
)

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

import chardet
import csv
from io import StringIO
import os


class TreeNodeViewSet(viewsets.ModelViewSet):
    """Base ViewSet for tree nodes with common functionality"""
    authentication_classes = [SessionAuthentication, BasicAuthentication]
    permission_classes = [IsAuthenticated]
    children_attr = None  # Must be set by child class
    children_serializer = None  # Must be set by child class
    select_related_fields = []
    prefetch_related_fields = []
    filter_fields = {}  # Map of query param to field lookup

    def get_queryset(self):
        """Get base queryset with proper select_related and prefetch_related"""
        queryset = super().get_queryset()
        
        # Apply base filters
        queryset = self.apply_base_filters(queryset)
        
        # Apply search filter
        queryset = self.apply_search_filter(queryset)
        
        # Apply specific filters
        queryset = self.apply_specific_filters(queryset)
        
        # Add select_related
        if self.select_related_fields:
            queryset = queryset.select_related(*self.select_related_fields)
            
        # Add prefetch_related only for detail actions to optimize list views
        if self.action != 'list' and self.prefetch_related_fields:
            queryset = queryset.prefetch_related(*self.prefetch_related_fields)
            
        return queryset.distinct()

    def apply_base_filters(self, queryset):
        """Apply any base filters (e.g., user access). Override in child classes."""
        return queryset

    def apply_search_filter(self, queryset):
        """Apply the general search filter"""
        filter_value = self.request.query_params.get('filter', '').strip()
        if filter_value:
            filter_conditions = self.get_search_conditions(filter_value)
            if filter_conditions:
                queryset = queryset.filter(filter_conditions)
        return queryset

    def get_search_conditions(self, filter_value):
        """Get Q objects for search. Override in child classes for custom search."""
        return Q(name__icontains=filter_value)

    def apply_specific_filters(self, queryset):
        """Apply specific field filters from filter_fields mapping"""
        for param, field in self.filter_fields.items():
            if value := self.request.query_params.get(param):
                # Handle multiple values for a single parameter
                if ',' in value:
                    values = [v.strip() for v in value.split(',')]
                    queryset = queryset.filter(**{f"{field}__in": values})
                else:
                    queryset = queryset.filter(**{field: value})
        return queryset

    @action(detail=True, methods=['get'])
    def children(self, request, pk=None):
        """Get children of a specific node with filtering and pagination"""
        if not self.children_attr or not self.children_serializer:
            return Response(
                {'error': 'Node does not support children'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Get parent node
            node = self.get_object()
            if not hasattr(node, self.children_attr):
                return Response(
                    {'error': f'Invalid children attribute: {self.children_attr}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get children queryset
            queryset = getattr(node, self.children_attr).all()

            # Apply filters
            filter_value = request.query_params.get('filter', '').strip()
            if filter_value:
                queryset = queryset.filter(name__icontains=filter_value)

            # Apply additional filters
            for param, field in self.filter_fields.items():
                if value := request.query_params.get(param):
                    if ',' in value:
                        values = [v.strip() for v in value.split(',')]
                        queryset = queryset.filter(**{f"{field}__in": values})
                    else:
                        queryset = queryset.filter(**{field: value})

            # Get pagination parameters
            limit = int(request.query_params.get('limit', 20))
            offset = int(request.query_params.get('offset', 0))
            
            # Get total count before slicing
            total = queryset.count()
            
            # Apply pagination
            children = queryset[offset:offset + limit + 1]
            has_more = len(children) > limit
            
            if has_more:
                children = children[:limit]
            
            # Serialize children
            serializer = self.children_serializer(children, many=True)
            
            return Response({
                'nodes': serializer.data,
                'has_more': has_more,
                'total': total,
                'offset': offset,
                'limit': limit
            })

        except Exception as e:
            return Response(
                {'error': f'Error fetching children: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def get_serializer_context(self):
        """Add request to serializer context"""
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def perform_create(self, serializer):
        """Add error handling to create"""
        try:
            return serializer.save()
        except ValidationError as e:
            raise e
        except Exception as e:
            raise ValidationError(str(e))

    def perform_update(self, serializer):
        """Add error handling to update"""
        try:
            return serializer.save()
        except ValidationError as e:
            raise e
        except Exception as e:
            raise ValidationError(str(e))


class ProjectAccessViewSet(viewsets.ModelViewSet):
    queryset = ProjectAccess.objects.all()
    serializer_class = ProjectAccessSerializer
    authentication_classes = [SessionAuthentication, BasicAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """
        Only return project access records for projects the user owns.
        """
        return ProjectAccess.objects.filter(project__owner=self.request.user)


class ProjectViewSet(TreeNodeViewSet):
    serializer_class = ProjectSerializer
    queryset = Project.objects.all()
    select_related_fields = ['owner']
    prefetch_related_fields = ['locations', 'user_access']
    children_attr = 'locations'
    children_serializer = LocationSerializer
    filter_fields = {
        'name': 'name__icontains',
        'type': 'project_type',
        'owner': 'owner_id'
    }

    def get_queryset(self):
        """
        Get queryset with filtering support.
        Special handling for the 'filter' parameter which searches multiple fields.
        """
        queryset = Project.objects.filter(
            Q(owner=self.request.user) |
            Q(user_access__user=self.request.user, 
              user_access__revoked_at__isnull=True)
        )

        # Apply general filter parameter
        if filter_value := self.request.query_params.get('filter', '').strip():
            queryset = queryset.filter(
                Q(name__icontains=filter_value) |
                Q(locations__name__icontains=filter_value)
            )

        # Apply other filters
        for param, field in self.filter_fields.items():
            if value := self.request.query_params.get(param):
                queryset = queryset.filter(**{field: value})

        # Apply select_related and prefetch_related
        if self.select_related_fields:
            queryset = queryset.select_related(*self.select_related_fields)
        
        if self.action != 'list' and self.prefetch_related_fields:
            queryset = queryset.prefetch_related(*self.prefetch_related_fields)

        return queryset.distinct()

    @action(detail=True, methods=['get'])
    def children(self, request, pk=None):
        """
        Get children (locations) for a project with filtering support.
        """
        project = self.get_object()
        queryset = project.locations.all()

        # Apply filters
        if filter_value := request.query_params.get('filter', '').strip():
            queryset = queryset.filter(name__icontains=filter_value)

        # Get pagination parameters
        limit = int(request.query_params.get('limit', 20))
        offset = int(request.query_params.get('offset', 0))

        # Get total count before slicing
        total = queryset.count()

        # Apply pagination
        queryset = queryset[offset:offset + limit + 1]
        has_more = len(queryset) > limit

        if has_more:
            queryset = queryset[:limit]

        serializer = self.children_serializer(queryset, many=True)

        return Response({
            'nodes': serializer.data,
            'has_more': has_more,
            'total': total,
            'offset': offset,
            'limit': limit
        })



class LocationViewSet(TreeNodeViewSet):
    serializer_class = LocationSerializer
    queryset = Location.objects.all()
    select_related_fields = ['project']
    prefetch_related_fields = ['measurements']
    children_attr = 'measurements'
    children_serializer = MeasurementSerializer
    filter_fields = {
        'name': 'name__icontains',
        'project': 'project_id'
    }

    def get_queryset(self):
        """
        Get queryset with filtering and access control.
        """
        queryset = Location.objects.filter(
            Q(project__owner=self.request.user) |
            Q(project__user_access__user=self.request.user,
              project__user_access__revoked_at__isnull=True)
        )

        # Apply general filter parameter
        if filter_value := self.request.query_params.get('filter', '').strip():
            queryset = queryset.filter(
                Q(name__icontains=filter_value) |
                Q(measurements__name__icontains=filter_value)
            )

        # Apply specific filters
        for param, field in self.filter_fields.items():
            if value := self.request.query_params.get(param):
                queryset = queryset.filter(**{field: value})

        # Apply select_related
        if self.select_related_fields:
            queryset = queryset.select_related(*self.select_related_fields)

        # Apply prefetch_related only for detail actions
        if self.action != 'list' and self.prefetch_related_fields:
            queryset = queryset.prefetch_related(*self.prefetch_related_fields)

        return queryset.distinct()

    @action(detail=True, methods=['get'])
    def children(self, request, pk=None):
        """
        Get children (measurements) for a location with filtering support.
        """
        location = self.get_object()
        queryset = location.measurements.all()

        # Apply filters
        if filter_value := request.query_params.get('filter', '').strip():
            queryset = queryset.filter(name__icontains=filter_value)

        # Get pagination parameters
        limit = int(request.query_params.get('limit', 20))
        offset = int(request.query_params.get('offset', 0))

        # Get total count before slicing
        total = queryset.count()

        # Apply pagination
        queryset = queryset[offset:offset + limit + 1]
        has_more = len(queryset) > limit

        if has_more:
            queryset = queryset[:limit]

        serializer = self.children_serializer(queryset, many=True)

        return Response({
            'nodes': serializer.data,
            'has_more': has_more,
            'total': total,
            'offset': offset,
            'limit': limit
        })



        # Existing implementation...

class DataCopyGrantViewSet(viewsets.ModelViewSet):
    queryset = DataCopyGrant.objects.all()
    serializer_class = DataCopyGrantSerializer
    authentication_classes = [SessionAuthentication, BasicAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """
        Only return grants where the user is involved.
        """
        user = self.request.user
        return DataCopyGrant.objects.filter(models.Q(from_user=user) | models.Q(to_user=user))

class DataImportViewSet(viewsets.ModelViewSet):
    serializer_class = DataImportSerializer
    queryset = DataImport.objects.all()

    def get_queryset(self):
        """
        Restrict access to the project owner or users with granted access.
        """
        user = self.request.user
        return DataImport.objects.filter(
            models.Q(dataset__data_source__project__owner=user) |
            models.Q(dataset__data_source__project__user_access__user=user, dataset__data_source__project__user_access__revoked_at__isnull=True)
        )


class DatasetViewSet(viewsets.ModelViewSet):
    serializer_class = DatasetSerializer
    queryset = Dataset.objects.all()

    def get_queryset(self):
        """
        Restrict dataset access to the project owner or users with granted access.
        """
        user = self.request.user
        return Dataset.objects.filter(
            models.Q(data_source__project__owner=user) |
            models.Q(data_source__project__user_access__user=user, data_source__project__user_access__revoked_at__isnull=True)
        )

class MeasurementCategoryViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = MeasurementCategorySerializer
    authentication_classes = [SessionAuthentication, BasicAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """
        Get queryset with optimized prefetching and optional filtering.
        """
        queryset = MeasurementCategory.objects.prefetch_related(
            'types',
            'types__units',
            Prefetch(
                'types__measurements',
                queryset=Measurement.objects.select_related('location', 'unit')
            )
        )

        # Filter by name if provided
        name = self.request.query_params.get('name')
        if name:
            queryset = queryset.filter(name__icontains=name)

        return queryset

    @action(detail=True, methods=['get'])
    def types(self, request, pk=None):
        """
        Get measurement types for a specific category.
        """
        category = self.get_object()
        types = category.types.prefetch_related('units').all()
        serializer = MeasurementTypeSerializer(types, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def measurements(self, request, pk=None):
        """
        Get all measurements in this category.
        """
        category = self.get_object()
        measurements = Measurement.objects.filter(
            type__category=category
        ).select_related(
            'type',
            'unit',
            'location',
            'location__project'
        )

        # Optional location filter
        location_id = request.query_params.get('location')
        if location_id:
            measurements = measurements.filter(location_id=location_id)

        serializer = MeasurementSerializer(measurements, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def stats(self, request, pk=None):
        """
        Get statistics about the category.
        """
        category = self.get_object()
        
        stats = {
            'type_count': category.types.count(),
            'measurement_count': Measurement.objects.filter(
                type__category=category
            ).count(),
            'unit_count': MeasurementUnit.objects.filter(
                type__category=category
            ).count(),
            'types': [{
                'name': t.name,
                'measurement_count': t.measurements.count(),
                'unit_count': t.units.count(),
            } for t in category.types.all()]
        }
        
        return Response(stats)


class MeasurementTypeViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = MeasurementTypeSerializer
    authentication_classes = [SessionAuthentication, BasicAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """
        Get queryset with optimized prefetching and optional filtering.
        """
        queryset = MeasurementType.objects.select_related(
            'category'
        ).prefetch_related(
            'units',
            Prefetch(
                'measurements',
                queryset=Measurement.objects.select_related('location', 'unit')
            )
        )

        # Filter by category if provided
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category_id=category)

        # Filter by name if provided
        name = self.request.query_params.get('name')
        if name:
            queryset = queryset.filter(name__icontains=name)

        # Filter by supports_multipliers if provided
        supports_multipliers = self.request.query_params.get('supports_multipliers')
        if supports_multipliers is not None:
            queryset = queryset.filter(supports_multipliers=supports_multipliers)

        return queryset

    @action(detail=True, methods=['get'])
    def units(self, request, pk=None):
        """
        Get units for a specific measurement type.
        """
        measurement_type = self.get_object()
        units = measurement_type.units.all()
        serializer = MeasurementUnitSerializer(units, many=True)
        
        # Include base unit information
        base_unit = next((unit for unit in units if unit.is_base_unit), None)
        response_data = {
            'units': serializer.data,
            'base_unit': MeasurementUnitSerializer(base_unit).data if base_unit else None
        }
        return Response(response_data)

    @action(detail=True, methods=['get'])
    def measurements(self, request, pk=None):
        """
        Get all measurements of this type.
        """
        measurement_type = self.get_object()
        measurements = measurement_type.measurements.select_related(
            'unit',
            'location',
            'location__project'
        ).all()

        # Optional location filter
        location_id = request.query_params.get('location')
        if location_id:
            measurements = measurements.filter(location_id=location_id)

        serializer = MeasurementSerializer(measurements, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def stats(self, request, pk=None):
        """
        Get statistics about the measurement type.
        """
        measurement_type = self.get_object()
        
        stats = {
            'measurement_count': measurement_type.measurements.count(),
            'unit_count': measurement_type.units.count(),
            'base_unit': next(
                (str(unit) for unit in measurement_type.units.all() if unit.is_base_unit),
                None
            ),
            'locations_count': measurement_type.measurements.values(
                'location'
            ).distinct().count(),
            'projects_count': measurement_type.measurements.values(
                'location__project'
            ).distinct().count()
        }
        
        return Response(stats)    

class MeasurementViewSet(TreeNodeViewSet):
    serializer_class = MeasurementSerializer
    queryset = Measurement.objects.all()
    select_related_fields = ['type', 'unit', 'location', 'location__project']
    children_attr = None  # Measurements are leaf nodes
    filter_fields = {
        'name': 'name__icontains',
        'location': 'location_id',
        'type': 'type_id',
        'category': 'type__category_id'
    }

    def get_queryset(self):
        """
        Get queryset with filtering and access control.
        """
        queryset = Measurement.objects.filter(
            Q(location__project__owner=self.request.user) |
            Q(location__project__user_access__user=self.request.user,
              location__project__user_access__revoked_at__isnull=True)
        )

        # Apply general filter parameter
        if filter_value := self.request.query_params.get('filter', '').strip():
            queryset = queryset.filter(name__icontains=filter_value)

        # Apply specific filters
        for param, field in self.filter_fields.items():
            if value := self.request.query_params.get(param):
                queryset = queryset.filter(**{field: value})

        # Apply select_related
        if self.select_related_fields:
            queryset = queryset.select_related(*self.select_related_fields)

        return queryset.distinct()

class ModelFieldsViewSet(viewsets.ViewSet):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    def list(self, request):
        """
        Get all field definitions and relationships for frontend form building.
        """
        # Get all measurement-related choices first
        categories = MeasurementCategory.objects.prefetch_related(
            'types',
            'types__units'
        ).all()

        measurement_choices = {
            'categories': [
                {
                    'id': cat.id,
                    'name': cat.name,
                    'display_name': cat.display_name,
                    'types': [
                        {
                            'id': type.id,
                            'name': type.name,
                            'description': type.description,
                            'supports_multipliers': type.supports_multipliers,
                            'units': [
                                {
                                    'id': unit.id,
                                    'name': unit.name,
                                    'is_base_unit': unit.is_base_unit,
                                    'conversion_factor': unit.conversion_factor
                                }
                                for unit in type.units.all()
                            ]
                        }
                        for type in cat.types.all()
                    ]
                }
                for cat in categories
            ],
            'multipliers': [
                {'id': choice[0], 'display_name': choice[1]}
                for choice in MeasurementType.MULTIPLIER_CHOICES
            ]
        }

        return Response({
            'project': {
                'level': 1,
                'model_name': 'project',
                'display_field': 'name',
                'fields': [
                    {
                        'name': 'name',
                        'type': 'string',
                        'required': True,
                        'display_field': True
                    },
                    {
                        'name': 'project_type',
                        'type': 'choice',
                        'required': True,
                        'choices': [
                            {'id': c[0], 'display_name': c[1]}
                            for c in Project.ProjectType
                        ]
                    },
                    {
                        'name': 'start_date',
                        'type': 'date',
                        'required': False
                    },
                    {
                        'name': 'end_date',
                        'type': 'date',
                        'required': False
                    }
                ],
                'child_type': 'location',
                'validation_rules': {
                    'end_date': ['after_start_date']
                }
            },
            'location': {
                'level': 2,
                'model_name': 'location',
                'display_field': 'name',
                'fields': [
                    {
                        'name': 'name',
                        'type': 'string',
                        'required': True,
                        'display_field': True
                    },
                    {
                        'name': 'address',
                        'type': 'text',
                        'required': True
                    },
                    {
                        'name': 'latitude',
                        'type': 'decimal',
                        'required': False,
                        'decimal_places': 6,
                        'max_digits': 9
                    },
                    {
                        'name': 'longitude',
                        'type': 'decimal',
                        'required': False,
                        'decimal_places': 6,
                        'max_digits': 9
                    }
                ],
                'parent_type': 'project',
                'child_type': 'measurement',
                'validation_rules': {
                    'latitude': ['valid_latitude'],
                    'longitude': ['valid_longitude']
                }
            },
            'measurement': {
                'level': 3,
                'model_name': 'measurement',
                'display_field': 'name',
                'fields': [
                    {
                        'name': 'name',
                        'type': 'string',
                        'required': True,
                        'display_field': True
                    },
                    {
                        'name': 'description',
                        'type': 'text',
                        'required': False
                    },
                    {
                        'name': 'category',
                        'type': 'foreign_key',
                        'required': True,
                        'choices': measurement_choices['categories']
                    },
                    {
                        'name': 'type',
                        'type': 'foreign_key',
                        'required': True,
                        'depends_on': 'category'
                    },
                    {
                        'name': 'unit',
                        'type': 'foreign_key',
                        'required': True,
                        'depends_on': 'type'
                    },
                    {
                        'name': 'multiplier',
                        'type': 'choice',
                        'required': False,
                        'choices': measurement_choices['multipliers'],
                        'conditional': 'type.supports_multipliers'
                    },
                    {
                        'name': 'source_timezone',
                        'type': 'choice',
                        'required': True,
                        'choices': [
                            {'id': tz, 'display_name': tz}
                            for tz in sorted(available_timezones())
                        ],
                        'default': 'UTC'
                    }
                ],
                'parent_type': 'location',
                'validation_rules': {
                    'unit': ['belongs_to_type'],
                    'multiplier': ['supported_by_type'],
                    'source_timezone': ['valid_timezone']
                }
            }
        })

    @action(detail=False, methods=['get'])
    def validation_rules(self, request):
        """
        Get all validation rules that can be applied to fields.
        """
        return Response({
            'after_start_date': {
                'type': 'date_comparison',
                'compare_to': 'start_date',
                'operator': 'gt'
            },
            'valid_latitude': {
                'type': 'range',
                'min': -90,
                'max': 90
            },
            'valid_longitude': {
                'type': 'range',
                'min': -180,
                'max': 180
            },
            'belongs_to_type': {
                'type': 'relationship',
                'check': 'unit.type_id === type_id'
            },
            'supported_by_type': {
                'type': 'conditional',
                'condition': 'type.supports_multipliers'
            },
            'valid_timezone': {
                'type': 'timezone'
            }
        })

    @action(detail=False, methods=['get'])
    def dependencies(self, request):
        """
        Get field dependencies for dynamic form updates.
        """
        return Response({
            'measurement': {
                'type': ['category'],
                'unit': ['type'],
                'multiplier': ['type.supports_multipliers']
            }
        })

@login_required(login_url='/')
def dashboard(request):
    """
    Serve the initial dashboard template.
    All data will be loaded dynamically via API endpoints.
    """
    try:
        context = {
            'model_fields': ModelFieldsSerializer(instance=None).to_representation(None),
            'initial_config': {
                'api_urls': {
                    'projects': reverse('project-list'),
                    'locations': reverse('location-list'),
                    'measurements': reverse('measurement-list'),
                }
            }
        }
        return render(request, 'main/dashboard.html', context)
        
    except Exception as e:
        print(f"Error in dashboard view: {str(e)}")
        return render(request, 'main/dashboard.html', {
            'error': 'An error occurred while loading the dashboard',
            'details': str(e)
        })

def excel_upload(request):
    """
    Handle upload of CSV files and convert to Excel with validation.
    """
    if 'csv_file' not in request.FILES:
        return JsonResponse({'error': 'CSV file is required'}, status=400)

    csv_file = request.FILES['csv_file']
    folder_path = request.POST.get('folder_path')
    workbook_name = request.POST.get('workbook_name')
    sheet_name = request.POST.get('sheet_name')

    if not all([folder_path, workbook_name, sheet_name]):
        return JsonResponse({'error': 'All fields are required'}, status=400)

    try:
        # Read and detect encoding
        raw_content = csv_file.read()
        detected_encoding = chardet.detect(raw_content)['encoding']
        file_content = raw_content.decode(detected_encoding)
        csv_reader = csv.reader(StringIO(file_content))
        rows = list(csv_reader)

        if not rows:
            return JsonResponse({'error': 'CSV file is empty'}, status=400)

        headers = rows[0]
        required_columns = ['category', 'type', 'unit']
        missing_columns = [col for col in required_columns if col not in headers]
        
        if missing_columns:
            return JsonResponse({
                'error': 'Missing required columns',
                'missing': missing_columns
            }, status=400)

        # Get column indices
        category_index = headers.index('category')
        type_index = headers.index('type')
        unit_index = headers.index('unit')

        # Validate against database
        categories = MeasurementCategory.objects.prefetch_related(
            'types__units'
        ).all()
        
        # Build validation lookup dictionaries
        valid_categories = {c.name.lower(): c for c in categories}
        valid_types = {}
        valid_units = {}
        
        for category in categories:
            category_types = {t.name.lower(): t for t in category.types.all()}
            valid_types[category.name.lower()] = category_types
            
            for type_obj in category.types.all():
                type_units = {u.name.lower(): u for u in type_obj.units.all()}
                valid_units[type_obj.name.lower()] = type_units

        # Validate rows
        validation_errors = []
        for row_num, row in enumerate(rows[1:], start=2):
            if not row or len(row) < max(category_index, type_index, unit_index) + 1:
                validation_errors.append(f'Row {row_num}: Invalid row length')
                continue

            category = row[category_index].strip().lower()
            meas_type = row[type_index].strip().lower()
            unit = row[unit_index].strip().lower()

            if not category:
                validation_errors.append(f'Row {row_num}: Empty category')
                continue

            if category not in valid_categories:
                validation_errors.append(
                    f'Row {row_num}: Invalid category "{row[category_index]}"'
                )
                continue

            if not meas_type:
                validation_errors.append(f'Row {row_num}: Empty type')
                continue

            if meas_type not in valid_types.get(category, {}):
                validation_errors.append(
                    f'Row {row_num}: Invalid type "{row[type_index]}" for category "{row[category_index]}"'
                )
                continue

            if not unit:
                validation_errors.append(f'Row {row_num}: Empty unit')
                continue

            if unit not in valid_units.get(meas_type, {}):
                validation_errors.append(
                    f'Row {row_num}: Invalid unit "{row[unit_index]}" for type "{row[type_index]}"'
                )

        if validation_errors:
            return JsonResponse({
                'error': 'Validation errors found',
                'errors': validation_errors
            }, status=400)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Add headers with validation lists
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            
            # Add data validation for known columns
            if header == 'category':
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(valid_categories.keys())}"',
                    allow_blank=False
                )
                ws.add_data_validation(dv)
                dv.add(f'{get_column_letter(col)}2:{get_column_letter(col)}1048576')
                
            elif header == 'type':
                # Type validation will be handled by custom Excel formulas
                pass
                
            elif header == 'unit':
                # Unit validation will be handled by custom Excel formulas
                pass

        # Add data rows
        for row in rows[1:]:
            ws.append(row)

        # Add dynamic validation formulas for types and units
        type_col = get_column_letter(type_index + 1)
        category_col = get_column_letter(category_index + 1)
        unit_col = get_column_letter(unit_index + 1)

        # Add helper sheets for lookups
        lookup_sheet = wb.create_sheet(title="Lookups")
        current_row = 1

        # Write category-type mappings
        for category, types in valid_types.items():
            lookup_sheet.cell(row=current_row, column=1, value=category)
            for col, type_name in enumerate(types.keys(), start=2):
                lookup_sheet.cell(row=current_row, column=col, value=type_name)
            current_row += 1

        # Write type-unit mappings
        unit_start_row = current_row + 1
        for type_name, units in valid_units.items():
            lookup_sheet.cell(row=current_row, column=1, value=type_name)
            for col, unit_name in enumerate(units.keys(), start=2):
                lookup_sheet.cell(row=current_row, column=col, value=unit_name)
            current_row += 1

        # Save workbook
        workbook_name = workbook_name if workbook_name.endswith('.xlsx') else f"{workbook_name}.xlsx"
        full_path = os.path.join(folder_path, workbook_name)
        wb.save(full_path)

        return JsonResponse({
            'status': 'success',
            'message': 'File processed successfully',
            'path': full_path,
            'stats': {
                'total_rows': len(rows) - 1,
                'categories': len(valid_categories),
                'types': sum(len(types) for types in valid_types.values()),
                'units': sum(len(units) for units in valid_units.values())
            }
        })

    except Exception as e:
        import traceback
        print(f"Error in excel_upload: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'error': 'An error occurred while processing the file',
            'details': str(e)
        }, status=500)

def index(request):
    """
    Serve the login page and handle authentication.
    """
    try:
        if request.method == 'POST':
            username = request.POST.get('username')
            password = request.POST.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('dashboard')
            else:
                return render(request, 'main/index.html', {'error': 'Invalid credentials'})
        return render(request, 'main/index.html')
    except Exception as e:
        print(f"Error in index view: {str(e)}")
        return render(request, 'main/index.html', {
            'error': 'An error occurred while processing login',
            'details': str(e)
        })

@login_required(login_url='/')
def project(request):
    """
    Serve the project detail view template.
    Dynamic data will be loaded via API endpoints.
    """
    try:
        context = {
            'model_fields': ModelFieldsSerializer(instance=None).to_representation(None),
            'project_types': dict(Project.ProjectType.choices),
            'api_urls': {
                'project': reverse('api-projects-list'),
                'project_access': reverse('api-project-access-list'),
                'export': reverse('api-projects-export')
            }
        }
        return render(request, 'main/project.html', context)
    except Exception as e:
        print(f"Error in project view: {str(e)}")
        return render(request, 'main/project.html', {
            'error': 'An error occurred while loading the project view',
            'details': str(e)
        })

@login_required(login_url='/')
def measurement(request):
    """
    Serve the measurement detail view template.
    Dynamic data will be loaded via API endpoints.
    """
    try:
        context = {
            'model_fields': ModelFieldsSerializer(instance=None).to_representation(None),
            'multiplier_choices': dict(MeasurementType.MULTIPLIER_CHOICES),
            'api_urls': {
                'measurements': reverse('api-measurements-list'),
                'categories': reverse('api-measurement-categories-list'),
                'types': reverse('api-measurement-types-list'),
                'units': reverse('api-measurement-units-list'),
                'timeseries': reverse('api-measurements-timeseries-list'),
                'export': reverse('api-measurements-export')
            }
        }
        
        # Add CodeMirror configuration if needed
        context['codemirror_config'] = {
            'mode': 'javascript',
            'theme': 'default',
            'lineNumbers': True
        }
        
        return render(request, 'main/measurement.html', context)
        
    except Exception as e:
        print(f"Error in measurement view: {str(e)}")
        return render(request, 'main/measurement.html', {
            'error': 'An error occurred while loading the measurement view',
            'details': str(e)
        })

@login_required(login_url='/')
def data(request):
    """
    Serve the data view template for data imports and visualization.
    Dynamic data will be loaded via API endpoints.
    """
    try:
        context = {
            'api_urls': {
                'imports': reverse('api-data-imports-list'),
                'datasets': reverse('api-datasets-list'),
                'data_sources': reverse('api-data-sources-list'),
                'preview': reverse('create-data-import'),
                'validate': reverse('excel-upload')
            },
            'upload_config': {
                'max_file_size': 50 * 1024 * 1024,  # 50MB
                'allowed_types': ['text/csv', 'application/vnd.ms-excel'],
                'allowed_extensions': ['.csv', '.txt']
            }
        }
        
        return render(request, 'main/data.html', context)
        
    except Exception as e:
        print(f"Error in data view: {str(e)}")
        return render(request, 'main/data.html', {
            'error': 'An error occurred while loading the data view',
            'details': str(e)
        })

@login_required(login_url='/')
def model(request):
    """
    Serve the model view template.
    Dynamic data will be loaded via API endpoints.
    """
    try:
        context = {
            'api_urls': {
                'categories': reverse('api-measurement-categories-list'),
                'types': reverse('api-measurement-types-list'),
                'units': reverse('api-measurement-units-list'),
                'export': reverse('api-model-export')
            },
            'model_fields': ModelFieldsSerializer(instance=None).to_representation(None)
        }
        return render(request, 'main/model.html', context)
    except Exception as e:
        print(f"Error in model view: {str(e)}")
        return render(request, 'main/model.html', {
            'error': 'An error occurred while loading the model view',
            'details': str(e)
        })

def get_preview_content(file_like_object, max_size=5000):
    """
    Get preview content from any file-like object
    Returns preview content and whether it was truncated
    """
    try:
        file_like_object.seek(0)
        # Read raw bytes first
        raw_content = file_like_object.read()
        if not raw_content:
            return "Empty file", False

        # Try UTF-8 first
        try:
            content = raw_content.decode('utf-8')
        except UnicodeDecodeError:
            # If UTF-8 fails, try to detect encoding
            import chardet
            detected = chardet.detect(raw_content)
            if detected and detected['encoding']:
                try:
                    content = raw_content.decode(detected['encoding'])
                except UnicodeDecodeError:
                    # If all else fails, try with 'latin1' (which always works but might be incorrect)
                    content = raw_content.decode('latin1')
            else:
                content = raw_content.decode('latin1')

        # Truncate if needed
        if len(content) > max_size:
            content = content[:max_size] + "\n... (content truncated for preview)"
            was_truncated = True
        else:
            was_truncated = False

        # Reset file pointer for future reads
        file_like_object.seek(0)
        return content, was_truncated

    except Exception as e:
        print(f"Error in get_preview_content: {str(e)}")
        raise ValueError(f'Error reading file content: {str(e)}')
    
@login_required
@require_http_methods(["POST"])
def create_data_import(request):
    """Create a new data import from a file upload"""
    try:
        # Validate inputs
        location_id = request.POST.get('location_id')
        if not location_id:
            return JsonResponse({'error': 'No location ID provided'}, status=400)

        # Get location and verify access
        try:
            location = Location.objects.select_related('project').get(id=location_id)
            project = location.project
            
            # Verify user has access to project
            if not (request.user == project.owner or 
                    ProjectAccess.objects.filter(
                        project=project,
                        user=request.user,
                        revoked_at__isnull=True
                    ).exists()):
                return JsonResponse({'error': 'No access to this location'}, status=403)
                
        except Location.DoesNotExist:
            return JsonResponse({'error': 'Invalid location ID'}, status=400)

        # Validate file
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)
        
        file = request.FILES['file']
        if not file.name.lower().endswith('.csv'):
            return JsonResponse({'error': 'Only CSV files are supported'}, status=400)

        # Create data source
        try:
            data_source, created = DataSource.objects.get_or_create(
                name=f'File Upload - {file.name}',
                project=project,
                defaults={
                    'source_type': 'file',
                    'description': f'File upload for location {location.name}',
                    'created_by': request.user
                }
            )
        except Exception as e:
            return JsonResponse({'error': f'Failed to create data source: {str(e)}'}, status=500)

        # Link data source to location
        try:
            DataSourceLocation.objects.get_or_create(
                data_source=data_source,
                location=location
            )
        except Exception as e:
            data_source.delete()  # Cleanup on failure
            return JsonResponse({'error': f'Failed to link data source to location: {str(e)}'}, status=500)

        # Create dataset
        try:
            dataset = Dataset.objects.create(
                data_source=data_source,
                name=f'Dataset - {file.name}',
                created_by=request.user
            )
        except Exception as e:
            data_source.delete()  # Cleanup on failure
            return JsonResponse({'error': f'Failed to create dataset: {str(e)}'}, status=500)

        # Create import record
        try:
            data_import = DataImport.objects.create(
                dataset=dataset,
                status='analyzing',
                created_by=request.user
            )
        except Exception as e:
            dataset.delete()  # Cleanup on failure
            return JsonResponse({'error': f'Failed to create import record: {str(e)}'}, status=500)

        # Generate preview
        try:
            preview_content, was_truncated = get_preview_content(file)
        except ValueError as e:
            # Clean up created objects on preview failure
            data_import.delete()
            dataset.delete()
            data_source.delete()
            return JsonResponse({'error': f'Failed to generate preview: {str(e)}'}, status=500)

        return JsonResponse({
            'import_id': data_import.id,
            'dataset_id': dataset.id,
            'preview_content': preview_content,
            'preview_truncated': was_truncated,
            'status': 'success'
        })

    except Exception as e:
        print(f"Error in create_data_import: {str(e)}")
        return JsonResponse({
            'error': str(e),
            'status': 'error'
        }, status=500)       


@login_required
@require_http_methods(["GET"])
def get_data_import(request, import_id):
    try:
        data_import = DataImport.objects.get(id=import_id)
        
        return JsonResponse({
            'id': data_import.id,
            'status': data_import.status,
            'created_at': data_import.started_at.isoformat(),
            'file_name': data_import.import_file.name if data_import.import_file else None,
        })
        
    except DataImport.DoesNotExist:
        return JsonResponse({'error': 'Import not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)